"""Extract a parity baseline from the source Excel workbook.

The workbook is the specification for `lihtc_screen`. This script dumps every
cached cell value Excel last computed into a JSON fixture so the Python engine
can be asserted against real Excel output, cell for cell, with no dependency on
Excel or LibreOffice at test time.

Usage:
    python tools/extract_baseline.py [workbook.xlsx] [out.json]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = REPO / "reference" / "Acq_Rehab_Model_v1.xlsx"
DEFAULT_OUT = REPO / "tests" / "fixtures" / "westbend_expected.json"


def _clean(value):
    """Normalise a cached cell value to something JSON-serialisable."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        # Excel stores everything as float; keep full precision.
        return float(value)
    return str(value)


def extract(workbook_path: Path) -> dict:
    values = openpyxl.load_workbook(workbook_path, data_only=True)
    formulas = openpyxl.load_workbook(workbook_path, data_only=False)

    sheets: dict[str, dict] = {}
    for ws_v in values.worksheets:
        ws_f = formulas[ws_v.title]
        cells: dict[str, object] = {}
        computed: dict[str, str] = {}
        for row in ws_f.iter_rows():
            for cell in row:
                raw = cell.value
                if raw is None:
                    continue
                if type(raw).__name__ == "ArrayFormula":
                    raw = raw.text
                coord = cell.coordinate
                cached = _clean(ws_v[coord].value)
                if cached is not None:
                    cells[coord] = cached
                if isinstance(raw, str) and raw.startswith("="):
                    computed[coord] = raw
        sheets[ws_v.title] = {
            "state": ws_v.sheet_state,
            "cells": cells,
            "formulas": computed,
        }

    return {
        "source": workbook_path.name,
        "sheet_order": [ws.title for ws in values.worksheets],
        "defined_names": {
            name: dn.value for name, dn in values.defined_names.items()
        },
        "sheets": sheets,
    }


def main() -> int:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not workbook.exists():
        print(f"workbook not found: {workbook}", file=sys.stderr)
        return 1

    baseline = extract(workbook)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(baseline, indent=1, sort_keys=True))

    n_cells = sum(len(s["cells"]) for s in baseline["sheets"].values())
    n_formulas = sum(len(s["formulas"]) for s in baseline["sheets"].values())
    print(f"{workbook.name}: {len(baseline['sheets'])} sheets, "
          f"{n_cells} cached values, {n_formulas} formulas -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
